import datetime
import time
import openpyxl
import random
import re
import threading
import pandas as pd
import yfinance as yf
from calendar import monthrange
import requests
import os

# warnings.filterwarnings("ignore", category=FutureWarning, module="yfinance")

start_date, end_date = datetime.datetime(1970, 2, 1), datetime.datetime.now()

# excel样式颜色
color = ['ffd2d2', 'ff8eff', 'd3a4ff', 'b9b9ff',
         'acd6ff', 'a6ffff', '4efeb3', '28ff28',
         'ff9d6f', 'ff9224', 'ffff37', '9aff02',
         'e1c4c4', 'dedebe', 'c4e1e1', 'e6e6f2']

pattern = [' download failed.', ' is blank.', ' is not comparable.', ' data odd.', r'retry over \d time.']


def main_process(Download=None, txt='failed_txt/failed.txt'):
    global failedList, old_failedList

    # 读取、清洗失败数据文本
    with open(txt, 'r', encoding='utf-8') as f:
        old_failedList = f.readlines()  # 存放下载失败数据
    failedList = list()

    for line in old_failedList:
        line = line.strip('\n')
        for p in pattern:
            line = re.sub(p, '', line)  # 将p去掉
        if line != '':
            failedList.append(line)

    if Download:
        Redownload(redownload_txt=txt)
    elif not Download:
        ToExcel()
    else:
        print('''
        Download = True    # 重新下载失败数据

        Download = False   # 将失败数据录入Excel
        ''')


fail_download = {'Shanghai_Shenzhen': [], 'Snp500_Ru1000': [], 'TSX': []}

def downloader(ticker, data_name, start_date, end_date, sleep_time=1, repeat=3, option=None):
    """
    下载股票数据，支持yfinance和requests两种方式，并可选择下载方式。
    参数:
        ticker (str): 股票代码。
        data_name (str): 数据类别名称（用于创建子文件夹）。
        start_date (datetime.date): 数据开始日期。
        end_date (datetime.date): 数据结束日期。
        sleep_time (int): 每次重试前的等待时间（秒）。
        repeat (int): 最大重试次数。
        option (int, optional): 下载方式选项。
            - None (或任何非0/1值): 先尝试yfinance，失败后回退到requests。
            - 0: 只尝试yfinance。
            - 1: 只尝试requests。
            默认为 None。
    """
    for _ in range(repeat):
        if option is None or option == 0:
            try:
                data = yf.Ticker(ticker).history(
                    period="max",
                    interval="1d",
                    start=start_date,
                    end=end_date,
                    prepost=False,
                    actions=False,
                    auto_adjust=False,
                    back_adjust=False,
                    proxy=None,
                    rounding=False
                )
                if data is None or data.shape[0] <= 1:
                    fail_download[data_name].append(ticker)
                    print(ticker + ' None')
                else:
                    # data = data.asfreq('ME', method='pad')  # 数据频度为每月（python环境3.11）
                    data = data.asfreq('M', method='pad')    # 数据频度为每月（python环境3.7）
                    data.to_csv(r'new_csv\\' + str(data_name) + r"\\" + str(ticker) + '.csv')
                    print(ticker + ' Successful')
                break
            except Exception as e:
                print(f"{ticker} Error: {e}")
                time.sleep(sleep_time)

        # 使用 requests 备用接口
        # 仅当 option 为 None (默认行为) 或 option 为 1 时尝试
        if option is None or option == 1:
            # 如果 option 为 0，在 yfinance 失败后应该跳过此部分。
            # 此检查确保如果 option 明确为 0 且 yfinance 失败，
            if option == 0 and 'e_yf' in locals(): # 检查 yfinance 是否失败且 option 是否为 0
                continue # 如果 option 为 0，则跳过此重试的 requests 尝试
            try:
                # 重新计算 yfinance 时间范围，确保结束日期是上个月的最后一天
                today = datetime.date.today()
                # 确定上个月的最后一天
                if today.month == 1:
                    # 如果当前月份是 1 月，则上个月是上一年的 12 月
                    target_month = 12
                    target_year = today.year - 1
                else:
                    target_month = today.month - 1
                    target_year = today.year
                target_day = monthrange(target_year, target_month)[1]
                req_end_date = datetime.date(target_year, target_month, target_day)
                start_unix = int(time.mktime(start_date.timetuple()))
                # Yahoo Finance API 的 end_unix 通常是独占的，所以加一天然后减去 1 秒
                end_unix = int(time.mktime((req_end_date + datetime.timedelta(days=1)).timetuple())) - 1
                url = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}?period1={start_unix}&period2={end_unix}&interval=1d&events=div%2Csplits"
                headers = {
                    "User-Agent": "Mozilla/5.0",
                    "Accept-Language": "en-US,en;q=0.9",
                    "Accept": "application/json",
                    "Referer": "https://finance.yahoo.com/",
                }
                response = requests.get(url, headers=headers, timeout=10)
                if response.status_code != 200:

                    raise ValueError(f"requests 状态码错误: {response.status_code}")
                result = response.json()
                chart = result.get("chart", {})
                if chart.get("error"):
                    fail_download[data_name].append(ticker)
                    raise ValueError(f"requests API 错误: {chart['error']}")

                result_data = chart.get("result", [])
                if not result_data:
                    fail_download[data_name].append(ticker)
                    raise ValueError("requests 未返回数据。")

                timestamps = result_data[0].get("timestamp", [])
                indicators = result_data[0].get("indicators", {})
                quote = indicators.get("quote", [{}])[0]
                adjclose_data = indicators.get("adjclose", [{}])[0]

                if not timestamps or not quote.get("close") or not adjclose_data.get("adjclose"):
                    fail_download[data_name].append(ticker)
                    raise ValueError("requests 数据缺失: 收盘价或调整收盘价数据缺失。")

                df = pd.DataFrame({
                    "Date": pd.to_datetime([datetime.datetime.utcfromtimestamp(ts) for ts in timestamps]),
                    "Open": quote.get("open", []),
                    "High": quote.get("high", []),
                    "Low": quote.get("low", []),
                    "Close": quote.get("close", []),
                    "Adj Close": adjclose_data.get("adjclose", []),
                    "Volume": quote.get("volume", []),
                })

                if df.shape[0] <= 1:
                    raise ValueError("requests 数据不足。")

                # 对齐到自然月的最后一天，填充缺失日期
                monthly_rows = []
                for ym, group in df.groupby(df['Date'].dt.to_period('M')):
                    last_day = ym.to_timestamp(how='end')
                    # 确保处理 'Close' 列中的 NaN 值以确定最后一个有效行
                    last_valid_row = group.loc[group['Close'].notna()].iloc[-1]
                    if last_valid_row['Date'].date() == last_day.date():
                        monthly_rows.append(last_valid_row)
                    else:
                        new_row = last_valid_row.copy()
                        new_row['Date'] = last_day
                        monthly_rows.append(new_row)

                df_monthly = pd.DataFrame(monthly_rows)
                df_monthly['Date'] = df_monthly['Date'].dt.strftime('%Y-%m-%d')

                os.makedirs(os.path.join('new_csv', data_name), exist_ok=True)
                filepath = os.path.join('new_csv', data_name, f"{ticker}.csv")
                df_monthly.to_csv(filepath, index=False)
                print(f"{ticker} requests API 下载成功。")
                return  # 成功下载，退出

            except Exception as e_req:
                if option == 1: # 如果只选择 requests，则不尝试 yfinance
                    print(f"{ticker} requests API 下载失败: {e_req}")
                    break # 退出重试循环
                else: # 如果 option 为 None，则返回 yfinance（通过外部循环重试）
                    print(f"{ticker} requests API 下载失败: {e_req}")
                    time.sleep(sleep_time)
                    # 如果 requests 失败且不是 requests-only 模式，它将重试外部循环，
                    # 外部循环将再次尝试 yfinance（如果 option 为 None）。

def Redownload(redownload_txt):
    print('正在重新下载失败数据...')

    name = ['Shanghai_Shenzhen', 'Snp500_Ru1000', 'TSX']

    global tickers

    record = ''
    # 获取下载列表
    thread_list = list()

    for n in range(0, 3):
        pos = failedList.index(name[n])
        try:
            failed_count = int(failedList[pos + 1].replace('失败下载数量: ', '').strip())
        except ValueError:
            print(f"Invalid format in failedList[pos + 1]: {failedList[pos + 1]}")
            failed_count = 0  # 默认值，或者根据需求处理

        download_tickers = failedList[(pos + 2): (pos + 2 + failed_count)]

        # 下载失败的重复5次，其余情况重复三次
        if redownload_txt == 'failed_txt/failed.txt':
            repeat = 5
        else:
            repeat = 3

        while repeat > 0:
            for ticker in download_tickers:
                # 确保Redownload也遵循下载方式选择
                t = threading.Thread(target=downloader, args=(ticker, name[n], start_date, end_date),
                                     kwargs={'option': download_option_method})
                thread_list.append(t)

            for i, thread in enumerate(thread_list):
                thread.start()  # 线程的设置
                if i % 3 == 0:
                    time.sleep(0.5)
                if i != 0 and i % 100 == 0:
                    time.sleep(10)
            for thread in thread_list:
                thread.join()

            thread_list.clear()
            repeat -= 1

            download_tickers = fail_download[name[n]]  # 更新下载失败的股票
            if repeat != 0:
                fail_download[name[n]] = []

        record += ('\n' + name[n] + '\n下载失败数量: ' + str(len(fail_download[name[n]])) + '\n' + '\n'.join(
            fail_download[name[n]]) + '\n')

    # 重新下载后仍下载失败的数据记录
    with open(redownload_txt, 'w', encoding='utf-8') as f2:
        f2.write(record)
    # 新旧数据对比
    with open(redownload_txt, 'r', encoding='utf-8') as f:
        new_failedList = f.readlines()

    compare_len = (len(old_failedList) - len(new_failedList))
    if compare_len <= 0:
        print('\n没有新数据，重试2-3次，然后录入Excel！')
    else:
        print('\n更新了 ' + str(compare_len) + ' 条数据！')

    print('重新下载完成！')


def symbolNo_download(symbolNo, download_option_method=None):  # v2.7: 添加 symbolNo_download 函数，用于单个股票的下载
    data = pd.ExcelFile('master_symbol_v1.6_2023.03.xlsx')
    sheet_names = data.sheet_names
    data_1 = pd.read_excel('master_symbol_v1.6_2023.03.xlsx', 2)  # Shanghai_Shenzhen
    data_1.name = sheet_names[2]
    data_2 = pd.read_excel('master_symbol_v1.6_2023.03.xlsx', 1)  # Snp500_Ru1000
    data_2.name = sheet_names[1]
    data_3 = pd.read_excel('master_symbol_v1.6_2023.03.xlsx', 3)  # TSX
    data_3.name = sheet_names[3]
    required_data = [data_1, data_2, data_3]

    classify = None
    # 查找symbolNo的股票种类
    for _data in required_data:
        for index, row in _data.iterrows():
            if symbolNo == row[0]:  # symbolNo在表格中
                if row['currently use'] != 'yes':  # 但不符合条件
                    print(symbolNo + ' 目前未在使用状态！')
                    return
                classify = _data.name  # 如果是，则完成分类
                break
        if classify != None:  # 用于停止循环
            break

    # 下载symbolNo股票
    downloader(symbolNo, classify, start_date, end_date, option=download_option_method)

def get_data(y, c1, c2):  # 获取待填入数据
    for cell in total_columns[y]:
        y_Tickers.append(cell.value)
    for cell in total_columns[c1]:
        content1.append(cell.value)
    if c2 == '':
        for cell in total_columns[7]: # 假设 c2 为空时，第 7 列（索引 6）用于空 content2
            content2.append('')
    else:
        for cell in total_columns[c2]:
            content2.append(cell.value)


def date_Date():  # 获取当前日期
    month = str(datetime.datetime.now().month)
    day = str(datetime.datetime.now().day)
    year = str(datetime.datetime.now().year)
    if len(month) == 1:
        month = '0' + month
    if len(day) == 1:
        day = '0' + day
    date = int(year + month + day)
    return date


def set_pos(name, sheet, fgcolor, cell3, cell4):  # 获取填入位置
    update_time = list()
    for cell in list(sheet.columns)[0]:
        update_time.append(cell.value)
    update_time.pop(0) # 移除标题行

    pos = failedList.index(name)  # name的位置
    num = int(failedList[pos + 1].replace('下载失败数量: ', ''))  # 下载失败的股票数量
    data = failedList[(pos + 2): (pos + 2 + num)]  # 下载失败的股票信息

    try:
        blank_pos = update_time.index(None) + 1 # 找到第一个空行
    except ValueError: # 如果没有空值，表示所有单元格都已填充，则追加到末尾
        blank_pos = len(update_time) + 1

    date = date_Date()

    for i in range(0, num):
        ticker_pos = y_Tickers.index(data[i])
        cell_pos = 'D' + str(blank_pos + i) # 日期列
        cell2_pos = 'A' + str(blank_pos + i) # 股票代码列
        cell3_pos = 'B' + str(blank_pos + i) # content1列
        cell4_pos = 'C' + str(blank_pos + i) # content2列

        sheet[cell_pos] = date
        # openpyxl.styles 模块需要导入 PatternFill
        from openpyxl.styles import PatternFill
        sheet[cell2_pos].fill = PatternFill(fill_type='solid', fgColor=fgcolor)
        sheet[cell2_pos] = data[i]
        sheet[cell3_pos] = cell3[ticker_pos]
        sheet[cell4_pos] = cell4[ticker_pos]


def ToExcel():
    print('正在录入Excel...')

    excel = openpyxl.load_workbook('master_symbol_v1.6_2023.03.xlsx')
    names = excel.sheetnames

    global total_columns, y_Tickers, content1, content2
    total_columns = list()
    y_Tickers = list()
    content1 = list()
    content2 = list()

    # 遍历工作表（假设1、2、3分别对应上海/深圳、标普500/罗素1000、多伦多证券交易所）
    for n in range(1, 4):
        fgcolor = color[random.randint(0, 13)]

        # 解析失败数据文本
        total_sheet = excel[names[n]]
        total_columns.clear() # 为每个工作表清除
        for column in total_sheet.columns:
            total_columns.append(column)

        y_Tickers.clear()
        content1.clear()
        content2.clear()

        # 根据您的Excel工作表结构调整列索引
        if n == 1:
            get_data(0, 1, 2)
        elif n == 2:
            get_data(1, 4, '')
        elif n == 3:
            get_data(1, 2, 5)

        # 假设 XXX Faillist 工作表在 names[n+5]
        sheet = excel[names[n + 5]]

        set_pos(names[n], sheet, fgcolor, content1, content2)


    excel.save('master_symbol_v1.6_2023.03.xlsx')
    print('\n' + '完成！')


# 用于下载方式选择的全局变量
download_option_method = None  # None: 先yfinance后requests; 0: 只yfinance; 1: 只requests

if __name__ == '__main__':
    # 您可以根据需要设置这些值
    symbolNo = None # 如果要下载单个股票，请在此处设置股票代码，例如 'AAPL'
    flag = True  # True: 重新下载失败数据; False: 将失败数据录入Excel
    download_option_method = 1 # 设置为0表示优先只使用yfinance进行所有下载 1为只使用requests下载 none为先yfinance然后requests下载

    if symbolNo is None:
        main_process(Download=flag, txt='failed_txt/failed.txt')
    else:
        symbolNo_download(symbolNo, download_option_method=download_option_method)