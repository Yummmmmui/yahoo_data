# coding: utf-8
import os
import openpyxl
import pandas as pd
import datetime
from util.time_design import time_justify

now = datetime.datetime.now()
end = datetime.datetime(now.year, now.month, 1) - datetime.timedelta(days=1)
endDate = time_justify()
if end.month >= 10:
    upDate = str(end.year) + "." + str(end.month)
else:
    upDate = str(end.year) + ".0" + str(end.month)

import pandas as pd

filename = 'master_symbol_v1.6_2023.03.xlsx'  # 请确保文件是 .xlsx 格式
data = pd.ExcelFile(filename, engine='openpyxl')  # 使用 openpyxl 引擎来读取 Excel 文件
sheet_names = data.sheet_names

# 读取不同的工作表
data_1 = pd.read_excel(filename, sheet_name=sheet_names[2], engine='openpyxl')  # Shanghai_Shenzhen
data_1.name = sheet_names[2]
data_2 = pd.read_excel(filename, sheet_name=sheet_names[1], engine='openpyxl')  # Snp500_Ru1000
data_2.name = sheet_names[1]
data_3 = pd.read_excel(filename, sheet_name=sheet_names[3], engine='openpyxl')  # TSX
data_3.name = sheet_names[3]


_data = [data_1, data_2, data_3]


def check(delete=False):
    """
    检查股票的日期更新情况。
    根据指定的国家列表检查每个国家的股票数据文件夹中的最新日期。
    如果某个股票的最新日期与指定的结束日期不一致，则将该股票标记为不符合月份条件。
    如果delete参数设置为True，则删除不符合条件的股票文件。

    参数：
        delete (bool, optional)：是否删除不符合条件的股票文件。默认为False。

    返回：
        None

    """
    # 检查股票的日期更新
    countries = ['Shanghai_Shenzhen', 'Snp500_Ru1000', 'TSX']
    source = r'new_csv\\'
    for country in countries:
        path = source + country # 国家股票数据文件夹的路径
        file = os.listdir(path) # 获取文件夹中的所有文件

        success = 0  # 下载成功的股票数量
        unproper = 0  # 不符合月份条件的股票数量

        print(country)
        fail_list = [] # 不符合条件的股票文件列表

        for filename in file:
            filename = os.path.join(path, filename)
            df = pd.read_csv(filename)
            date = df.values[-1][0]
            success = success + 1
            if date != endDate:  # 如果最新日期与指定的结束日期不一致
                fail_list.append(filename)
                unproper = unproper + 1

        # 最后，删除不符合条件的股票
        if delete:
            for filename in fail_list:
                os.remove(filename)
        print("下载成功的股票数量：" + str(success))
        print("不符合月份条件的股票数量：" + str(unproper))


def sum():
    """
    汇总股票数据的相关信息到总结表格中。
    检查股票数据文件夹中每个国家的股票数量，并统计下载的股票数量和符合月份条件的股票数量。
    将这些信息汇总到总结表格中，并保存为Excel文件。

    """

    filename = openpyxl.load_workbook("QC_report_" + upDate + ".xlsx")
    if "Summary_cnt" not in filename.sheetnames:
        filename.create_sheet("Summary_cnt", 0)
        filename.save("QC_report_" + upDate + ".xlsx")

    s = filename['Summary_cnt'] # 获取Summary_cnt工作表
    s['A1'] = "country"
    s['B1'] = "tickers of master_sheet"
    s['C1'] = "threshold"
    s['D1'] = "total downloaded"
    s['E1'] = endDate
    print(endDate)  # 调试输出，确保它是正确的日期

    filename.save("QC_report_" + upDate + ".xlsx")

    countries = ['Shanghai_Shenzhen', 'Snp500_Ru1000', 'TSX']
    source = r'new_csv\\' # 股票数据文件夹的路径
    n = 0

    for country in countries:
        t_mus = 0 # t_mus为所需下载的股票总数量
        for index, row in _data[n].iterrows():
            if row['currently use'] == 'yes':
                t_mus = t_mus + 1

        path = source + country
        file = os.listdir(path)
        # t_down为总共下载的股票数量
        t_down = 0
        # dow_yes为下载到正确月份的股票数量
        dow_yes = 0

        for filename in file:
            filename = os.path.join(path, filename)
            df = pd.read_csv(filename)
            date = df.values[-1][0]
            # 获取最后一行的日期，并将其转换为日期格式
            date = pd.to_datetime(df.values[-1][0]).strftime('%Y-%m-%d')  # 转换为字符串格式 "YYYY-MM-DD"
            t_down = t_down + 1

            if date == endDate:
                dow_yes = dow_yes + 1

        filename = openpyxl.load_workbook("QC_report_" + upDate + ".xlsx")
        sheet = filename["Summary_cnt"]
        sheet['A' + str(n + 2)] = country # 设置国家单元格的值
        sheet['B' + str(n + 2)] = t_mus # 设置所需下载的股票总数量单元格的值
        sheet['C' + str(n + 2)] = int(0.9 * t_mus) # 设置阈值单元格的值
        sheet['D' + str(n + 2)] = t_down # 设置总共下载的股票数量单元格的值
        sheet['E' + str(n + 2)] = dow_yes # 设置下载到正确月份的股票数量单元格的值
        n = n + 1
        sheet['C4'] = 220
        filename.save("QC_report_" + upDate + ".xlsx")

flag = False

if __name__ == '__main__':
    if flag:
        # 最后删除不符合月份条件的股票时，将delete值改为True
        check(delete=False)
    else:
        # 每个月首次运行sum时，生成总结文件Qc_tabl e.xlsx
        sum()
